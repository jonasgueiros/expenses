import sqlite3
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from db_conn import conectar, fechar_conexao
from openpyxl import load_workbook
import os  # Import os to handle file paths

def criar_tabela(conexao, table_name, columns):
    """Cria uma tabela no banco de dados."""
    try:
        sql_create_table = f"CREATE TABLE IF NOT EXISTS {table_name} ({columns});"
        conexao.execute(sql_create_table)
        messagebox.showinfo("Success", f"Tabela '{table_name}' criada ou já existe.")
    except sqlite3.Error as e:
        messagebox.showerror("Error", f"Erro ao criar a tabela: {e}")

def deletar_tabela(conexao, table_name):
    """Deleta uma tabela do banco de dados."""
    try:
        sql_delete_table = f"DROP TABLE IF EXISTS {table_name};"
        conexao.execute(sql_delete_table)
        messagebox.showinfo("Success", f"Tabela '{table_name}' deletada com sucesso.")
    except sqlite3.Error as e:
        messagebox.showerror("Error", f"Erro ao deletar a tabela: {e}")

def atualizar_lista_tabelas(tree):
    """Atualiza a lista de tabelas na interface."""
    conexao = conectar()
    if conexao:
        cursor = conexao.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT IN ('sqlite_sequence', 'expenses');")
        tables = cursor.fetchall()
        tree.delete(*tree.get_children())  # Clear existing entries
        for table in tables:
            tree.insert("", "end", values=(table[0],))
        fechar_conexao(conexao)

def executar_com_verificacao(funcao, *args):
    """Executa uma função com verificação de conexão."""
    conexao = conectar()
    if conexao:
        funcao(conexao, *args)
        fechar_conexao(conexao)

def importar_excel(conexao):
    """Importa dados de um arquivo Excel e salva no banco de dados."""
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if not file_path:
        return  # User canceled the file dialog

    table_name = os.path.splitext(os.path.basename(file_path))[0]

    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao ler o arquivo Excel: {e}")
        return

    columns = []
    for cell in sheet[1]:  # Assuming the first row contains the headers
        column_name = cell.value
        if isinstance(column_name, str):
            column_name = column_name.replace(" ", "_")  # Replace spaces with underscores
        columns.append(f"{column_name} TEXT")  # Default to TEXT type

    columns_str = ", ".join(columns)
    criar_tabela(conexao, table_name, columns_str)

    try:
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
            values = [None if v is None else v for v in row]  # Handle None values
            placeholders = ', '.join(['?'] * len(values))
            sql_insert = f"INSERT INTO {table_name} VALUES ({placeholders})"
            conexao.execute(sql_insert, values)
        conexao.commit()  # Commit the changes
        messagebox.showinfo("Success", f"Dados importados para a tabela '{table_name}' com sucesso.")
    except Exception as e:
        messagebox.showerror("Error", f"Erro ao inserir dados na tabela: {e}")

def modificar_tabela_interface(table_name):
    def on_save_changes():
        new_table_name = new_table_name_entry.get()
        column_updates = []
        for entry in column_entries:
            col_name = entry[0].get()
            col_type = entry[1].get()
            if col_name and col_type:
                column_updates.append(f"{col_name} {col_type}")
        if new_table_name:
            executar_com_verificacao(modificar_tabela, table_name, new_table_name, column_updates)
            window.destroy()  # Close the modification window
        else:
            messagebox.showwarning("Input Error", "Por favor, preencha todos os campos.")

    window = tk.Toplevel()
    window.title("Modificar Tabela")

    # Table name entry
    tk.Label(window, text="Novo nome da tabela:").grid(row=0, column=0)
    new_table_name_entry = tk.Entry(window)
    new_table_name_entry.insert(0, table_name)  # Pre-fill the table name
    new_table_name_entry.grid(row=0, column=1)

    # Column entries
    column_entries = []
    conexao = conectar()
    if conexao:
        cursor = conexao.cursor()
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns = cursor.fetchall()
        for i, column in enumerate(columns):
            if column[1] not in ['datahoje', 'id']:  # Skip 'datahoje' and 'id' columns
                tk.Label(window, text=f"Coluna {i+1}:").grid(row=i+1, column=0)
                col_name_entry = tk.Entry(window)
                col_name_entry.insert(0, column[1])  # Pre-fill the column name
                col_name_entry.grid(row=i+1, column=1)
                tk.Label(window, text="Tipo de dados:").grid(row=i+1, column=2)
                col_type_entry = tk.Entry(window)
                col_type_entry.insert(0, column[2])  # Pre-fill the column type
                col_type_entry.grid(row=i+1, column=3)
                column_entries.append((col_name_entry, col_type_entry))
        fechar_conexao(conexao)

    tk.Button(window, text="Salvar alterações", command=on_save_changes).grid(row=len(columns)+1, column=0, columnspan=4)

    window.mainloop()

def modificar_tabela(conexao, old_table_name, new_table_name, column_updates):
    """Modifica a tabela no banco de dados."""
    try:
        # Step 1: Create a new table with the new structure
        columns_definition = ", ".join(column_updates)
        conexao.execute(f"CREATE TABLE {new_table_name} ({columns_definition});")

        # Step 2: Copy data from the old table to the new table
        columns_names = ", ".join([update.split()[0] for update in column_updates])
        conexao.execute(f"INSERT INTO {new_table_name} ({columns_names}) SELECT {columns_names} FROM {old_table_name};")

        # Step 3: Drop the old table
        conexao.execute(f"DROP TABLE {old_table_name};")

        # Step 4: Rename the new table to the old table's name
        conexao.execute(f"ALTER TABLE {new_table_name} RENAME TO {old_table_name};")

        conexao.commit()  # Commit the changes
        messagebox.showinfo("Success", f"Tabela '{old_table_name}' modificada com sucesso.")
    except sqlite3.Error as e:
        messagebox.showerror("Error", f"Erro ao modificar a tabela: {e}")

def criar_tabela_interface():
    def on_create_table():
        # Open the creation table window
        criar_tabela_window()

    def on_delete_table():
        selected_item = table_tree.selection()
        if selected_item:
            table_name = table_tree.item(selected_item)['values'][0]
            if table_name == 'expenses':
                messagebox.showwarning("Deletion Error", "Não é possível deletar a tabela 'expenses'.")
            else:
                if messagebox.askyesno("Confirmação", f"Você tem certeza que deseja deletar a tabela '{table_name}'?"):
                    executar_com_verificacao(deletar_tabela, table_name)
                    atualizar_lista_tabelas(table_tree)  # Refresh the table list
        else:
            messagebox.showwarning("Selection Error", "Por favor, selecione uma tabela para deletar.")

    def on_import_excel():
        conexao = conectar()
        if conexao:
            importar_excel(conexao)
            fechar_conexao(conexao)

    def on_modify_table():
        selected_item = table_tree.selection()
        if selected_item:
            table_name = table_tree.item(selected_item)['values'][0]
            modificar_tabela_interface(table_name)
        else:
            messagebox.showwarning("Selection Error", "Por favor, selecione uma tabela para modificar.")

    root = tk.Tk()
    root.title("Gerenciador de Tabelas")

    root.geometry("310x230+0+650")  # Width x Height + X Offset + Y Offset
    root.resizable(False, False)   # Disable resizing

    # Button frame
    button_frame = tk.Frame(root)
    button_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=10)

    tk.Button(button_frame, text="Criar Tabela", command=on_create_table).pack(side=tk.TOP, fill=tk.X)
    tk.Button(button_frame, text="Importar", command=on_import_excel).pack(side=tk.TOP, fill=tk.X)
    tk.Button(button_frame, text="Modificar", command=on_modify_table).pack(side=tk.TOP, fill=tk.X)
    tk.Button(button_frame, text="Deletar", command=on_delete_table).pack(side=tk.TOP, fill=tk.X)

    # Table list
    table_tree = ttk.Treeview(root, columns=("table_name",), show="headings")
    table_tree.heading("table_name", text="Nome da Tabela")
    table_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    atualizar_lista_tabelas(table_tree)  # Initialize the table list

    root.mainloop()
    
def criar_tabela_window():
    # Create a new window for creating a table
    window = tk.Toplevel()
    window.title("Criar Tabela")

    # Table name entry
    tk.Label(window, text="Nome da Tabela:").grid(row=0, column=0)
    table_name_entry = tk.Entry(window)
    table_name_entry.grid(row=0, column=1)

    # Data type entries
    data_type_entries = {
        "TEXT": tk.Entry(window),
        "INTEGER": tk.Entry(window),
        "REAL": tk.Entry(window),
        "BLOB": tk.Entry(window)
    }

    # Create labels and entry fields for each data type
    row = 1
    for data_type, entry in data_type_entries.items():
        tk.Label(window, text=f"{data_type}:").grid(row=row, column=0)
        entry.grid(row=row, column=1)
        row += 1
        
    def on_create_table():
        table_name = table_name_entry.get()
        columns = "id INTEGER PRIMARY KEY AUTOINCREMENT, datahoje TEXT"
        for data_type, entry in data_type_entries.items():
            column_names = entry.get().strip()
            if column_names:  # Only add if there are column names
                for name in column_names.split(','):
                    columns += f", {name.strip()} {data_type}"
        if table_name:
            executar_com_verificacao(criar_tabela, table_name, columns)
            window.destroy()  # Close the creation window
        else:
            messagebox.showwarning("Input Error", "Por favor, preencha todos os campos.")

    tk.Button(window, text="Criar Tabela", command=on_create_table).grid(row=row, column=0, columnspan=2)

    window.mainloop()

if __name__ == "__main__":
    criar_tabela_interface()