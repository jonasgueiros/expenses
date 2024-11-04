import flet as ft
import os
import sqlite3
from db_conn import conectar, fechar_conexao
from openpyxl import load_workbook

class NovaTab(ft.UserControl):
    def __init__(self, page):
        super().__init__()
        self.page = page
        self.create_widgets()

    def build(self):
        return ft.Container(
            content=ft.Column(
                controls=[
                    ft.Text("Criar Nova Tabela", size=20),
                    ft.ElevatedButton("Create Table", on_click=self.create_table_window),
                    ft.ElevatedButton("Importar Tabela do Excel", on_click=self.importar_excel)
                ]
            )
        )

    def create_widgets(self):
        pass

    def create_table_window(self, e):
        def on_create_table(e):
            table_name = table_name_entry.value
            columns = "id INTEGER PRIMARY KEY AUTOINCREMENT, datahoje TEXT"
            for data_type, entry in data_type_entries.items():
                column_names = entry.value.strip() if entry.value else ""
                if column_names:  # Only add if there are column names
                    for name in column_names.split(','):
                        columns += f", {name.strip()} {data_type}"
            if table_name:
                self.executar_com_verificacao(self.criar_tabela, table_name, columns)
                criar_tabela_dialog.open = False  # Close the creation window
                if self.page:
                    self.page.update()
            else:
                dialog = ft.AlertDialog(title=ft.Text("Input Error"), content=ft.Text("Por favor, preencha todos os campos."))
                if self.page:
                    self.page.overlay.append(dialog)
                    dialog.open = True

        table_name_entry = ft.TextField(label="Nome da Tabela")

        data_type_entries = {
            "TEXT": ft.TextField(label="TEXT"),
            "INTEGER": ft.TextField(label="INTEGER"),
            "REAL": ft.TextField(label="REAL"),
            "BLOB": ft.TextField(label="BLOB")
        }

        criar_tabela_dialog = ft.AlertDialog(
            title=ft.Text("Criar Tabela"),
            content=ft.Column(
                controls=[
                    table_name_entry,
                    *[entry for entry in data_type_entries.values()],
                    ft.ElevatedButton("Criar Tabela", on_click=on_create_table)
                ]
            ),
            actions=[ft.TextButton("Fechar", on_click=lambda e: setattr(criar_tabela_dialog, 'open', False))]
        )

        if self.page:
            self.page.overlay.append(criar_tabela_dialog)
            criar_tabela_dialog.open = True
            self.page.update()

    def criar_tabela(self, conexao, table_name, columns):
        """Cria uma tabela no banco de dados."""
        try:
            sql_create_table = f"CREATE TABLE IF NOT EXISTS {table_name} ({columns});"
            conexao.execute(sql_create_table)
            dialog = ft.AlertDialog(title=ft.Text("Success"), content=ft.Text(f"Tabela '{table_name}' criada ou já existe."))
            if self.page:
                self.page.overlay.append(dialog)
                dialog.open = True
        except sqlite3.Error as e:
            dialog = ft.AlertDialog(title=ft.Text("Error"), content=ft.Text(f"Erro ao criar a tabela: {e}"))
            if self.page:
                self.page.overlay.append(dialog)
                dialog.open = True

    def importar_excel(self, e):
        file_picker = ft.FilePicker(on_result=self.on_file_picked)
        if self.page:
            self.page.overlay.append(file_picker)
        file_picker.pick_files(allow_multiple=False, file_type=ft.FilePickerFileType.CUSTOM, allowed_extensions=["xlsx", "xls"])

    def on_file_picked(self, e):
        file_path = e.files[0].path if e.files else None
        if not file_path:
            return  # User canceled the file dialog

        table_name = os.path.splitext(os.path.basename(file_path))[0]

        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
        except Exception as e:
            dialog = ft.AlertDialog(title=ft.Text("Erro"), content=ft.Text(f"Erro ao ler o arquivo Excel: {e}"))
            if self.page:
                self.page.overlay.append(dialog)
                dialog.open = True
            return

        conexao = conectar()
        if conexao:
            try:
                if sheet is not None:
                    columns = ", ".join([f"{cell.value} TEXT" for cell in sheet[1]])
                else:
                    dialog = ft.AlertDialog(title=ft.Text("Erro"), content=ft.Text("A planilha está vazia ou não pôde ser lida."))
                    if self.page:
                        self.page.overlay.append(dialog)
                        dialog.open = True
                    return
                sql_create_table = f"CREATE TABLE IF NOT EXISTS {table_name} ({columns});"
                conexao.execute(sql_create_table)

                if sheet is not None:
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        placeholders = ", ".join(["?" for _ in row])
                        sql_insert = f"INSERT INTO {table_name} VALUES ({placeholders})"
                        conexao.execute(sql_insert, row)

                conexao.commit()
                dialog = ft.AlertDialog(title=ft.Text("Success"), content=ft.Text(f"Table '{table_name}' imported successfully."))
                if self.page:
                    self.page.overlay.append(dialog)
                    dialog.open = True
            except sqlite3.Error as e:
                dialog = ft.AlertDialog(title=ft.Text("Error"), content=ft.Text(f"Error importing table: {e}"))
                if self.page:
                    self.page.overlay.append(dialog)
                    dialog.open = True
            finally:
                fechar_conexao(conexao)

    def executar_com_verificacao(self, func, *args):
        conexao = conectar()
        if conexao:
            try:
                func(conexao, *args)
            finally:
                fechar_conexao(conexao)
